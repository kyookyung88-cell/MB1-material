import streamlit as st
import pandas as pd
import plotly.express as px
import plotly.graph_objects as go
from openpyxl import load_workbook
from datetime import datetime, date
import os
import json

# ── 페이지 설정 ──
st.set_page_config(
    page_title="MB1팀 소재 제안 및 채택 현황",
    page_icon="🧪",
    layout="wide",
)

# ── 파일 경로 ──
EXCEL_FILE = os.path.join(os.path.dirname(__file__), "MB1팀 소재 제안 및 채택 현황 (Ver1.0, 251203).xlsx")
DATA_FILE = os.path.join(os.path.dirname(__file__), "data_store.json")

# ── 컬럼 정의 (2026 시트 기준) ──
COLUMNS = [
    "날짜", "고객사", "요청사항/컨셉", "베네핏", "NO", "소재명", "INCI", "효능",
    "Story", "특허", "중국", "EWG", "비건", "RTB", "임상",
    "Recommended dose", "Clinical dose", "자사코드", "채택여부", "고객사 반응", "담당자"
]


def load_excel_data():
    """엑셀 파일에서 2026 시트 데이터를 파싱"""
    wb = load_workbook(EXCEL_FILE, data_only=True)
    ws = wb["2026"]

    # 병합셀 값 전파를 위한 매핑
    merged_values = {}
    for merged_range in ws.merged_cells.ranges:
        min_row = merged_range.min_row
        min_col = merged_range.min_col
        val = ws.cell(min_row, min_col).value
        for row in range(merged_range.min_row, merged_range.max_row + 1):
            for col in range(merged_range.min_col, merged_range.max_col + 1):
                merged_values[(row, col)] = val

    rows = []
    for row_idx in range(3, ws.max_row + 1):
        row_data = {}
        for col_idx in range(1, 22):  # A~U (21 columns)
            cell_val = ws.cell(row_idx, col_idx).value
            if cell_val is None and (row_idx, col_idx) in merged_values:
                cell_val = merged_values[(row_idx, col_idx)]
            col_name = COLUMNS[col_idx - 1] if col_idx - 1 < len(COLUMNS) else f"Col{col_idx}"
            row_data[col_name] = cell_val

        # 소재명이 있는 행만 유효 데이터
        if row_data.get("소재명"):
            # 날짜 처리
            if isinstance(row_data.get("날짜"), datetime):
                row_data["날짜"] = row_data["날짜"].strftime("%Y-%m-%d")
            elif row_data.get("날짜") is None:
                row_data["날짜"] = ""
            rows.append(row_data)

    return rows


def load_all_data():
    """엑셀 + JSON 저장소 데이터를 합쳐 반환"""
    excel_rows = load_excel_data()
    extra_rows = []
    if os.path.exists(DATA_FILE):
        with open(DATA_FILE, "r", encoding="utf-8") as f:
            extra_rows = json.load(f)
    all_rows = excel_rows + extra_rows
    df = pd.DataFrame(all_rows, columns=COLUMNS)
    return df


def save_new_entries(entries: list):
    """새 데이터를 JSON 저장소에 추가 (여러 건)"""
    existing = []
    if os.path.exists(DATA_FILE):
        with open(DATA_FILE, "r", encoding="utf-8") as f:
            existing = json.load(f)
    existing.extend(entries)
    with open(DATA_FILE, "w", encoding="utf-8") as f:
        json.dump(existing, f, ensure_ascii=False, indent=2)


# ── 데이터 로드 ──
df = load_all_data()

# ── 사이드바: 필터 ──
st.sidebar.title("필터")

# 담당자 필터
담당자_list = sorted(df["담당자"].dropna().unique().tolist())
selected_담당자 = st.sidebar.multiselect("담당자", 담당자_list, default=담당자_list)

# 고객사 필터
고객사_list = sorted(df["고객사"].dropna().unique().tolist())
selected_고객사 = st.sidebar.multiselect("고객사", 고객사_list, default=고객사_list)

# 채택여부 필터
채택_options = ["전체", "채택 (O)", "미채택", "미정"]
selected_채택 = st.sidebar.radio("채택여부", 채택_options)

# 필터 적용
filtered = df.copy()
if selected_담당자:
    filtered = filtered[filtered["담당자"].isin(selected_담당자) | filtered["담당자"].isna()]
if selected_고객사:
    filtered = filtered[filtered["고객사"].isin(selected_고객사) | filtered["고객사"].isna()]
if selected_채택 == "채택 (O)":
    filtered = filtered[filtered["채택여부"] == "O"]
elif selected_채택 == "미채택":
    filtered = filtered[filtered["채택여부"].isin(["X", "N"])]
elif selected_채택 == "미정":
    filtered = filtered[filtered["채택여부"].isna() | (filtered["채택여부"] == "")]

# ── 메인 타이틀 ──
st.title("MB1팀 소재 제안 및 채택 현황 Dashboard")

# ── 탭 구성 ──
tab1, tab2, tab3 = st.tabs(["📊 현황 대시보드", "📋 데이터 조회", "➕ 데이터 입력"])

# ══════════════════════════════════════════
# 탭1: 현황 대시보드
# ══════════════════════════════════════════
with tab1:
    # KPI 카드
    col1, col2, col3, col4 = st.columns(4)
    total = len(filtered)
    adopted = len(filtered[filtered["채택여부"] == "O"])
    rate = (adopted / total * 100) if total > 0 else 0
    unique_clients = filtered["고객사"].dropna().nunique()

    col1.metric("총 제안 소재 수", f"{total}건")
    col2.metric("채택 소재 수", f"{adopted}건")
    col3.metric("채택률", f"{rate:.1f}%")
    col4.metric("고객사 수", f"{unique_clients}개")

    st.divider()

    # 차트 영역
    chart_col1, chart_col2 = st.columns(2)

    with chart_col1:
        st.subheader("고객사별 제안 현황")
        client_counts = filtered.groupby("고객사").size().reset_index(name="건수")
        client_counts = client_counts.sort_values("건수", ascending=True)
        if not client_counts.empty:
            fig = px.bar(client_counts, x="건수", y="고객사", orientation="h",
                         color="건수", color_continuous_scale="Blues")
            fig.update_layout(height=400, showlegend=False)
            st.plotly_chart(fig, use_container_width=True)

    with chart_col2:
        st.subheader("채택여부 분포")
        채택_map = filtered["채택여부"].fillna("미정").replace("", "미정")
        채택_counts = 채택_map.value_counts().reset_index()
        채택_counts.columns = ["채택여부", "건수"]
        if not 채택_counts.empty:
            fig = px.pie(채택_counts, values="건수", names="채택여부",
                         color_discrete_sequence=px.colors.qualitative.Set2)
            fig.update_layout(height=400)
            st.plotly_chart(fig, use_container_width=True)

    chart_col3, chart_col4 = st.columns(2)

    with chart_col3:
        st.subheader("베네핏별 제안 현황")
        benefit_series = filtered["베네핏"].dropna().str.split(r"[,&\n]", regex=True).explode().str.strip()
        benefit_series = benefit_series[benefit_series != ""]
        if not benefit_series.empty:
            benefit_counts = benefit_series.value_counts().head(15).reset_index()
            benefit_counts.columns = ["베네핏", "건수"]
            fig = px.bar(benefit_counts, x="베네핏", y="건수",
                         color="건수", color_continuous_scale="Greens")
            fig.update_layout(height=400, showlegend=False, xaxis_tickangle=-45)
            st.plotly_chart(fig, use_container_width=True)

    with chart_col4:
        st.subheader("효능 키워드 TOP 15")
        efficacy_series = filtered["효능"].dropna().str.split(r"[,\n]", regex=True).explode().str.strip()
        efficacy_series = efficacy_series[efficacy_series != ""]
        if not efficacy_series.empty:
            efficacy_counts = efficacy_series.value_counts().head(15).reset_index()
            efficacy_counts.columns = ["효능", "건수"]
            fig = px.bar(efficacy_counts, x="효능", y="건수",
                         color="건수", color_continuous_scale="Oranges")
            fig.update_layout(height=400, showlegend=False, xaxis_tickangle=-45)
            st.plotly_chart(fig, use_container_width=True)

    # 담당자별 현황
    st.subheader("담당자별 제안 및 채택 현황")
    담당자_df = filtered[filtered["담당자"].notna()].copy()
    if not 담당자_df.empty:
        summary = 담당자_df.groupby("담당자").agg(
            제안건수=("소재명", "count"),
            채택건수=("채택여부", lambda x: (x == "O").sum())
        ).reset_index()
        summary["채택률(%)"] = (summary["채택건수"] / summary["제안건수"] * 100).round(1)
        fig = go.Figure()
        fig.add_trace(go.Bar(name="제안건수", x=summary["담당자"], y=summary["제안건수"],
                             marker_color="#636EFA"))
        fig.add_trace(go.Bar(name="채택건수", x=summary["담당자"], y=summary["채택건수"],
                             marker_color="#00CC96"))
        fig.update_layout(barmode="group", height=350)
        st.plotly_chart(fig, use_container_width=True)

    # 월별 추이
    st.subheader("월별 제안 추이")
    dated = filtered[filtered["날짜"].notna() & (filtered["날짜"] != "")].copy()
    if not dated.empty:
        dated["월"] = pd.to_datetime(dated["날짜"], errors="coerce").dt.to_period("M").astype(str)
        monthly = dated.groupby("월").size().reset_index(name="건수")
        fig = px.line(monthly, x="월", y="건수", markers=True)
        fig.update_layout(height=300)
        st.plotly_chart(fig, use_container_width=True)

    # 특허/중국/EWG/비건 보유 현황
    st.subheader("인증 및 특허 보유 현황")
    cert_cols = ["특허", "중국", "EWG", "비건"]
    cert_data = []
    for col_name in cert_cols:
        has = filtered[col_name].isin(["O", "Y", "등록 중"]).sum() if col_name in filtered.columns else 0
        cert_data.append({"항목": col_name, "보유": has, "미보유": total - has})
    cert_df = pd.DataFrame(cert_data)
    fig = go.Figure()
    fig.add_trace(go.Bar(name="보유", x=cert_df["항목"], y=cert_df["보유"], marker_color="#00CC96"))
    fig.add_trace(go.Bar(name="미보유", x=cert_df["항목"], y=cert_df["미보유"], marker_color="#EF553B"))
    fig.update_layout(barmode="stack", height=300)
    st.plotly_chart(fig, use_container_width=True)


# ══════════════════════════════════════════
# 탭2: 데이터 조회
# ══════════════════════════════════════════
with tab2:
    st.subheader("전체 데이터 조회")

    # 검색
    search = st.text_input("소재명/INCI/효능 검색", "")
    display_df = filtered.copy()
    if search:
        mask = (
            display_df["소재명"].fillna("").str.contains(search, case=False, na=False)
            | display_df["INCI"].fillna("").str.contains(search, case=False, na=False)
            | display_df["효능"].fillna("").str.contains(search, case=False, na=False)
        )
        display_df = display_df[mask]

    st.info(f"총 {len(display_df)}건 조회됨")

    # 주요 컬럼만 표시
    display_cols = ["날짜", "고객사", "베네핏", "소재명", "INCI", "효능", "특허", "중국",
                    "EWG", "비건", "채택여부", "담당자"]
    st.dataframe(
        display_df[display_cols],
        use_container_width=True,
        height=600,
    )

    # CSV 다운로드
    csv = filtered.to_csv(index=False).encode("utf-8-sig")
    st.download_button("CSV 다운로드", csv, "MB1_소재현황.csv", "text/csv")


# ══════════════════════════════════════════
# 탭3: 데이터 입력
# ══════════════════════════════════════════
with tab3:
    st.subheader("새 소재 제안 입력")
    st.caption("공통 정보를 한 번 입력하고, 소재를 여러 개 추가할 수 있습니다.")

    # ── 세션 초기화 ──
    if "materials" not in st.session_state:
        st.session_state.materials = [{}]  # 소재 1개로 시작

    # ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
    # STEP 1: 공통 정보 (고객사 단위)
    # ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
    st.markdown("### 1. 공통 정보")
    cc1, cc2, cc3, cc4 = st.columns(4)
    with cc1:
        common_date = st.date_input("날짜", value=date.today(), key="common_date")
    with cc2:
        common_client = st.text_input("고객사", key="common_client")
    with cc3:
        common_concept = st.text_input("요청사항 / 컨셉", key="common_concept")
    with cc4:
        common_manager = st.text_input("담당자", key="common_manager")

    cc5, cc6 = st.columns(2)
    with cc5:
        common_benefit = st.text_input("베네핏", key="common_benefit")
    with cc6:
        common_reaction = st.text_input("고객사 반응", key="common_reaction")

    st.divider()

    # ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
    # STEP 2: 소재 목록 (동적 추가/삭제)
    # ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
    st.markdown("### 2. 소재 목록")

    btn_col1, btn_col2, _ = st.columns([1, 1, 4])
    with btn_col1:
        if st.button("+ 소재 추가", type="primary", use_container_width=True):
            st.session_state.materials.append({})
            st.rerun()
    with btn_col2:
        if len(st.session_state.materials) > 1:
            if st.button("- 마지막 삭제", use_container_width=True):
                st.session_state.materials.pop()
                st.rerun()

    # 각 소재 입력 블록
    for i in range(len(st.session_state.materials)):
        with st.expander(f"소재 {i + 1}", expanded=True):
            r1c1, r1c2, r1c3 = st.columns(3)
            with r1c1:
                st.text_input("소재명 *", key=f"m_name_{i}")
                st.text_input("INCI", key=f"m_inci_{i}")
            with r1c2:
                st.text_input("효능", key=f"m_efficacy_{i}")
                st.text_area("Story", height=68, key=f"m_story_{i}")
            with r1c3:
                st.text_input("특허 (O/X/등록 중)", key=f"m_patent_{i}")
                st.text_input("RTB", key=f"m_rtb_{i}")

            r2c1, r2c2, r2c3, r2c4, r2c5, r2c6 = st.columns(6)
            with r2c1:
                st.selectbox("중국", ["", "O", "X"], key=f"m_china_{i}")
            with r2c2:
                st.selectbox("EWG", ["", "O", "X"], key=f"m_ewg_{i}")
            with r2c3:
                st.selectbox("비건", ["", "O", "X"], key=f"m_vegan_{i}")
            with r2c4:
                st.text_input("임상", key=f"m_clinical_{i}")
            with r2c5:
                st.text_input("Rec. dose", key=f"m_rec_dose_{i}")
            with r2c6:
                st.text_input("Clin. dose", key=f"m_clin_dose_{i}")

            r3c1, r3c2 = st.columns(2)
            with r3c1:
                st.text_input("자사코드", key=f"m_code_{i}")
            with r3c2:
                st.selectbox("채택여부", ["", "O", "X"], key=f"m_adopted_{i}")

    st.divider()

    # ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
    # STEP 3: 저장
    # ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
    if st.button("전체 저장", type="primary", use_container_width=True):
        entries = []
        errors = []
        for i in range(len(st.session_state.materials)):
            name = st.session_state.get(f"m_name_{i}", "").strip()
            if not name:
                errors.append(f"소재 {i + 1}: 소재명이 비어 있습니다.")
                continue
            entries.append({
                "날짜": common_date.strftime("%Y-%m-%d"),
                "고객사": common_client or None,
                "요청사항/컨셉": common_concept or None,
                "베네핏": common_benefit or None,
                "NO": i + 1,
                "소재명": name,
                "INCI": st.session_state.get(f"m_inci_{i}", "") or None,
                "효능": st.session_state.get(f"m_efficacy_{i}", "") or None,
                "Story": st.session_state.get(f"m_story_{i}", "") or None,
                "특허": st.session_state.get(f"m_patent_{i}", "") or None,
                "중국": st.session_state.get(f"m_china_{i}", "") or None,
                "EWG": st.session_state.get(f"m_ewg_{i}", "") or None,
                "비건": st.session_state.get(f"m_vegan_{i}", "") or None,
                "RTB": st.session_state.get(f"m_rtb_{i}", "") or None,
                "임상": st.session_state.get(f"m_clinical_{i}", "") or None,
                "Recommended dose": st.session_state.get(f"m_rec_dose_{i}", "") or None,
                "Clinical dose": st.session_state.get(f"m_clin_dose_{i}", "") or None,
                "자사코드": st.session_state.get(f"m_code_{i}", "") or None,
                "채택여부": st.session_state.get(f"m_adopted_{i}", "") or None,
                "고객사 반응": common_reaction or None,
                "담당자": common_manager or None,
            })

        if errors:
            for e in errors:
                st.warning(e)
        if entries:
            save_new_entries(entries)
            st.success(f"{len(entries)}건의 소재가 저장되었습니다!")
            st.session_state.materials = [{}]
            st.rerun()
        elif not errors:
            st.error("저장할 소재가 없습니다. 소재명을 입력해주세요.")

    # ── 추가된 데이터 관리 ──
    if os.path.exists(DATA_FILE):
        with open(DATA_FILE, "r", encoding="utf-8") as f:
            extra = json.load(f)
        if extra:
            st.divider()
            st.subheader("추가 입력된 데이터")
            extra_df = pd.DataFrame(extra)
            st.dataframe(extra_df, use_container_width=True)

            if st.button("추가 데이터 전체 삭제", type="secondary"):
                os.remove(DATA_FILE)
                st.success("추가 데이터가 삭제되었습니다.")
                st.rerun()
